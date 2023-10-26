from flask import Flask, send_file,url_for,render_template,request,redirect,flash,jsonify,send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_mail import Mail, Message
import os,datetime
from datetime import datetime
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = "abc"  
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///bank.db'

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'ubbbankindia@gmail.com'
app.config['MAIL_PASSWORD'] = 'wliirdcdkisptpxs'  # Replace with your actual application-specific password


UPLOAD_FOLDER = 'sts'  # Define the folder where you want to save the files
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


mail = Mail(app)

db = SQLAlchemy(app)
migrate = Migrate(app,db)


class User(db.Model):
    cust_id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), nullable=False)
    phone = db.Column(db.Integer, nullable=False)
    email = db.Column(db.String(80), nullable=False)
    address = db.Column(db.String(80), nullable=False)
    initial_dep = db.Column(db.String(80), nullable=False)
    def __init__(self,name,phone,email,address,initial_dep):
        self.name=name
        self.phone=phone
        self.email =email
        self.address=address
        self.initial_dep = initial_dep
class admin(db.Model):
    emp_id = db.Column(db.Integer, primary_key=True)
    password=db.Column(db.String(10),nullable=False)
class AC(db.Model):
    ac_no = db.Column(db.Integer, primary_key=True)
    cust_id = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(80), nullable=False)
    phone = db.Column(db.Integer, nullable=False)
    def __init__(self,name,phone,cust_id):
        self.name=name
        self.phone=phone
        self.cust_id=cust_id
@app.route('/',methods=["POST","GET"])
def welcome():
    if request.method == "POST":
        a = admin.query.all()
        for item in a:
            emp_id = item.emp_id
            password2 = item.password
        u_name = int(request.form["username"])
        password1 = request.form["password"]
        if u_name == emp_id and password2 == password1:
            return render_template("stafficon.html")
        else:
            return"sorry"
    return render_template("welcome.html")
@app.route('/update',methods=["POST","GET"])
def update():
    if request.method == "POST":
        add_cus = User(request.form["name"],request.form["mb"],request.form["email"],request.form["address"],request.form["initial_dep"])
        dep = request.form["initial_dep"]
        a = request.form["email"]
        db.session.add(add_cus)
        db.session.commit()
        specific_users = User.query.filter_by(email=request.form["email"]).all()
        for item in specific_users:
            cust_id = item.cust_id
        acc_num=AC(name=request.form["name"],phone=request.form["mb"],cust_id=cust_id)
        db.session.add(acc_num)
        db.session.commit()
        specific_users = AC.query.filter_by(phone=request.form["mb"]).all()
        for item in specific_users:
            ac_no = item.ac_no
        current_date = datetime.now().strftime("%Y-%m-%d")
        file_name = os.path.join(app.config['UPLOAD_FOLDER'] , f'Customer{cust_id}.xlsx')
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['ACCOUNT NO:',str(ac_no), 'NAME :',request.form["name"], 'ACCOUNT OPEN DATE :',current_date])
        sheet.append(['Date', 'Transaction Type', 'Amount'])
        workbook.save(file_name)
        # f = open(file_name, "w")
        # f.write("ACCOUNT NO: " + str(ac_no) + "      cust_id: " + str(cust_id) + "      open date: " + current_date + "\n" + current_date + "      DEPOSIT"+"         "+str(dep) + "\n")
        
        b = "Your account has opened successfully your acountc no is:"+str(ac_no)+"and your initial deposit amount is "+str(dep)
        c = "UBB BANK"
        d = "AC OPENED SUCCESSFULLY"
        return redirect(url_for('sms', a=a, b=b, c=c , d=d))
    return render_template("staffhome.html")
@app.route('/deposit', methods=["POST", "GET"])
def deposit():
    if request.method == "POST":
        ac_no = int(request.form["ac_no"])
        dep_amt = int(request.form["amount"])
        specific_users = AC.query.filter_by(ac_no=ac_no).all()  # Use 'ac_no' instead of 'ac_no'
        for item in specific_users:
            cust_id = item.cust_id
        record = User.query.get(cust_id)
        a = record.email
        record.initial_dep = int(record.initial_dep)+dep_amt
        db.session.commit()
        c = "UBB BANK"
        b = "Your account number" + str(ac_no) +"has been deposited with" + str(dep_amt) + " your current balance is"+ str(record.initial_dep)
        d = "amount deposited"
        current_date = datetime.now().strftime("%Y-%m-%d")
        file_name = os.path.join(app.config['UPLOAD_FOLDER'] , f'Customer{cust_id}.xlsx')
        workbook = load_workbook(file_name)
        sheet = workbook.active
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.append([timestamp, "DEPOSIT", dep_amt])
        workbook.save(file_name) 
        return redirect(url_for('sms', a=a, b=b, c=c,d=d))
    return render_template("deposit.html")
@app.route('/withdraw', methods=["POST", "GET"])
def withdraw():
    if request.method == "POST":
        Acc = int(request.form["ac_no"])
        amt = int(request.form["amount"])
        record = AC.query.get(Acc)
        cus = record.cust_id
        record1 = User.query.get(cus)
        bal = record1.initial_dep
        if amt>int(bal):
            flash("INSUFFICIENT BALANCE")
            return render_template("stafficon.html")
        else:
            record1.initial_dep = int(record1.initial_dep) - amt
            db.session.commit()
            current_date = datetime.now().strftime("%Y-%m-%d")
            file_name = os.path.join(app.config['UPLOAD_FOLDER'] , f'Customer{cus}.xlsx')
            workbook = load_workbook(file_name)
            sheet = workbook.active
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            sheet.append([timestamp, "WITHDRAW", amt])
            workbook.save(file_name) 
            # f = open(file_name, "a")
            # f.write(  current_date + "      WITHDRAW"+"                   "+str(amt) + "\n")
        
            a = record1.email
            c = "UBB BANK"
            b = "Your account number" + str(Acc)+"has been debited with" + str(amt) + " your current balance is"+ str(record1.initial_dep)
            d = "amount debited"
            return redirect(url_for('sms', a=a, b=b, c=c,d=d))
        
        
    return render_template("withdraw.html")
@app.route('/transfer', methods=["POST", "GET"])
def transfer():
    if request.method == "POST":
        fac = int(request.form["fac"])
        tac = int(request.form["tac"])
        amt = int(request.form["amt"])
        record = AC.query.get(fac)
        record1 = AC.query.get(tac)
        cus = record.cust_id
        cus1 = record1.cust_id
        fetchac = User.query.get(cus)
        fetchac.initial_dep = int(fetchac.initial_dep)-amt
        db.session.commit()
        fetchac1 = User.query.get(cus1)
        x = fetchac.initial_dep
        msg = Message( "UBB BANK", sender='ubbbankindia@gmail.com', recipients=[fetchac.email])
        msg.body = "Amount "+ str(amt)+" is transfered from your account " +str(fac)+ "to "+ fetchac1.name+"("+str(tac)+")"+"avl_bal"+str(x)
        mail.send(msg)
        current_date = datetime.now().strftime("%Y-%m-%d")
        file_name = os.path.join(app.config['UPLOAD_FOLDER'] , f'Customer{cus}.xlsx')
        workbook = load_workbook(file_name)
        sheet = workbook.active
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.append([timestamp, "TFR TO :"+str(tac)+" "+fetchac1.name, amt])
        workbook.save(file_name) 
        # f = open(file_name, "a")
        # f.write(  current_date + "      TFR TO "+"\n"+str(tac)+fetchac1.name+"                   "+str(amt) )
        fetchac1.initial_dep = int(fetchac1.initial_dep)+amt
        db.session.commit()
        x = fetchac1.initial_dep
        msg = Message( "UBB BANK", sender='ubbbankindia@gmail.com', recipients=[fetchac1.email])
        msg.body = "Amount "+ str(amt)+" is received from " +fetchac.name+"("+str(fac)+")"+"avl_bal"+str(x)
        mail.send(msg)
        file_name = os.path.join(app.config['UPLOAD_FOLDER'] , f'Customer{cus1}.xlsx')
        workbook = load_workbook(file_name)
        sheet = workbook.active
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet.append([timestamp, "TFR FROM :"+str(fac)+" "+fetchac.name, amt])
        workbook.save(file_name) 
        # f = open(file_name, "a")
        # f.write(  current_date + "      TFR FROM "+"\n"+str(fac)+fetchac.name+"                   "+str(amt) )
        return "done"
    return render_template("transfer.html")

@app.route('/sms', methods=["GET","POST"])
def sms():
    if request.method == "GET":
        a = request.args.get('a')
        b = request.args.get('b')
        c = request.args.get('c')
        d = request.args.get('d')
        msg = Message( c, sender='ubbbankindia@gmail.com', recipients=[a])
        msg.body = b
        mail.send(msg)
        flash(d)
        return redirect(url_for("demo"))
@app.route('/api/getname', methods=['GET'])
def get_name():
    id = int(request.args.get('ac'))
    record = AC.query.get(id)
    if not record:
        a = "A/C NOT FOUND"
        return jsonify({'name': a})
    cid = record.cust_id
    record2 = User.query.get(cid)
    x = record2.name
    a = x.upper()
    return jsonify({'name': a})
@app.route('/getdetails', methods=["GET","POST"])
def get_details():
    if request.method == "POST":
        ac = request.form["ac_no"]
        record = AC.query.get(ac)
        if not record:
            a = "A/C NOT FOUND"
            flash(a)
            return render_template("getdetails.html")
        cid = record.cust_id
        record2 = User.query.get(cid)
        return render_template("displaydetails.html",data=record2,ac=ac)
    return render_template("getdetails.html")
@app.route("/download/<int:cus1>")
def download(cus1):
    try:
        directory = app.config['UPLOAD_FOLDER']
        f = f"Customer{cus1}.xlsx"
        return send_from_directory(directory, f , as_attachment=True)
    except:
        flash("STS NOT FOUND")
        return render_template("stafficon.html")


if __name__ ==('__main__'):
    with app.app_context():
        db.create_all()
    app.run(debug='True')